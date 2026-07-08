"""Generate FAIR_assessment_checklist.xlsx.

Scores the deposit against the RDA FAIR Data Maturity Model core indicators
applicable to a clinical dataset. 15 core indicators: 13 fully met (1.0) and 2
partially met (0.5) - I1 (semantic interoperability) and R1.3 (long-term
preservation). Each row carries an evidence statement and a reference to the
deposit file/metadata field supporting the score.

Output is written to the deposit root as FAIR_assessment_checklist.xlsx.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "FAIR_assessment_checklist.xlsx"

# (Principle, Indicator ID, Indicator text, Score, Evidence, Deposit reference)
ROWS = [
    ("Findable", "F1", "Data assigned a globally unique and persistent identifier", 1.0,
     "Dataset has a Zenodo DOI (10.5281/zenodo.20301027), a globally unique persistent identifier.",
     "DOI 10.5281/zenodo.20301027; README.md citation"),
    ("Findable", "F2", "Data described with rich metadata", 1.0,
     "Full variable-level metadata provided for every file and field.",
     "data/codebook_en.csv; data/codebook_en.md"),
    ("Findable", "F3", "Metadata clearly and explicitly include the identifier of the data", 1.0,
     "Zenodo record metadata and README embed the DOI alongside the described files.",
     "README.md; Zenodo record"),
    ("Findable", "F4", "Data registered/indexed in a searchable resource", 1.0,
     "Deposited in Zenodo and indexed by DataCite, making the record discoverable.",
     "Zenodo / DataCite index"),
    ("Accessible", "A1", "Data retrievable by identifier using a standardised protocol", 1.0,
     "Record and files retrievable over HTTPS via the resolved DOI.",
     "Zenodo HTTPS; DOI resolver"),
    ("Accessible", "A1.1", "Protocol is open, free and universally implementable", 1.0,
     "HTTPS is open, free and universally implementable.",
     "Zenodo access protocol"),
    ("Accessible", "A2", "Metadata accessible even when the data are no longer available", 1.0,
     "Zenodo retains DOI-linked metadata (tombstone) independent of file availability.",
     "Zenodo preservation policy"),
    ("Interoperable", "I1",
     "Data use a formal, accessible, shared, broadly applicable knowledge representation (semantic interoperability)",
     0.5,
     "Infection sites and pathogens map to a controlled internal vocabulary with cross-references to NHSN standard terminology; full SNOMED-CT / ICD-10-CM coding not yet implemented (resource constraints).",
     "code/translation_maps.py; data/codebook_en.csv"),
    ("Interoperable", "I2", "Data use vocabularies that themselves follow FAIR principles", 1.0,
     "Categorical values normalised to a documented, versioned English vocabulary released with the deposit.",
     "code/translation_maps.py"),
    ("Interoperable", "I3", "Data include qualified references to other (meta)data", 1.0,
     "Codebook links each variable to its file and to the department/site/pathogen crosswalks.",
     "data/codebook_en.csv; data/codebook_en.md"),
    ("Reusable", "R1", "Data richly described with a plurality of accurate and relevant attributes", 1.0,
     "57 documented variables per episode including temporal, departmental, microbiological, device and outcome attributes.",
     "data/cases_long_en.csv; data/codebook_en.csv"),
    ("Reusable", "R1.1", "(Meta)data released with a clear and accessible data usage license", 1.0,
     "Code released under MIT; data released under CC-BY 4.0, both stated explicitly.",
     "LICENSE; README.md"),
    ("Reusable", "R1.2", "(Meta)data associated with detailed provenance", 1.0,
     "Provenance keys (YYYY-MM source_file) and a documented six-stage pipeline describe data origin and processing.",
     "code/run_all.py; code/stage1_ingestion.py..stage6_deidentification_qa.py"),
    ("Reusable", "R1.3",
     "(Meta)data meet domain-relevant community standards (long-term preservation)",
     0.5,
     "Zenodo guarantees 10-year retention with DOI persistence; indefinite archival (e.g., national archive) not yet established.",
     "Zenodo retention policy"),
    ("Reusable", "R1.4", "Data deposited in a trusted, certified repository", 1.0,
     "Deposited in Zenodo, a CERN-backed general-purpose research repository.",
     "Zenodo record"),
]

wb = Workbook()
ws = wb.active
ws.title = "FAIR assessment"

headers = ["Principle", "Indicator ID", "Indicator", "Score (0/0.5/1.0)",
           "Status", "Evidence", "Deposit reference"]
ws.append(headers)
hdr_fill = PatternFill("solid", fgColor="305496")
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = hdr_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)

full = partial = 0
for principle, iid, text, score, evidence, ref in ROWS:
    status = "Fully met" if score == 1.0 else "Partially met"
    if score == 1.0:
        full += 1
    else:
        partial += 1
    ws.append([principle, iid, text, score, status, evidence, ref])

# Summary row.
ws.append([])
ws.append(["SUMMARY", "", f"{len(ROWS)} core indicators",
           round(sum(r[3] for r in ROWS) / len(ROWS), 3),
           f"{full} fully met, {partial} partially met", "", ""])

widths = [14, 12, 46, 16, 16, 60, 44]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = w
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
ws.freeze_panes = "A2"

wb.save(OUT)
print(f"wrote {OUT.name}: {full} fully met + {partial} partial of {len(ROWS)} indicators")
